import xlrd
from openpyxl import load_workbook

# 从指定行、列开始读Excel 数据
def read_info( xlsxfile, startRow, startCol):
    info = []
    try:
        fr = xlrd.open_workbook(xlsxfile)
    except Exception:
        return 0
    table = fr.sheets()[0]
    cols = table.ncols
    rows = table.nrows
    for m in range(startRow - 1, rows):
        a = ""
        for n in range(startCol - 1, cols):
            b = table.cell(m, n).value
            a = a + str(b)
            if n <cols:
                a = a + ','
        info.append(a)
    return info

# 读取指定行列单元格
def read_cell(xlsxfile, Row,Col) :
    if Row<=0:
        print("Row should >=1")
        return "error"
    elif Col <=0:
        print("Col should >=1")
        return "error"
    else :
        try :
            fr = xlrd.open_workbook(xlsxfile)
        except Exception :
            return 0
        table = fr.sheets()[0]
        return table.cell(Row-1,Col-1).value

# 从指定行、列开始写入文件
def write_info(filename, textlist, startRow, startCol):
    wb=load_workbook(filename)
    sheet = wb[wb.get_sheet_names()[0]]
    info = []
    for result in textlist:
        info.append(result.split(','))
    for i in range(0,len(info)):
        for j in range(len(info[i])):
            sheet.cell(row = startRow + i,column= startCol + j).value = str(info[i][j])
    wb.save(filename)

# 写入指定单元格
def write_cell(filename, value, Row, Col):
    if Row<=0:
        print("Row should >=1")
        return "error"
    elif Col <=0:
        print("Col should >=1")
        return "error"
    else:
        wb=load_workbook(filename)
        sheet = wb[wb.get_sheet_names()[0]]
        sheet.cell(row = Row,column= Col).value = str(value)
        wb.save(filename)

# 从最后一行开始写入新数据
def write_result(filename, textlist):
    wb=load_workbook(filename)
    sheet = wb[wb.get_sheet_names()[0]]
    row = sheet.max_row+1
    case = []
    for result in textlist:
        case.append(str(result).split(','))
    for i in range(len(case)):
        newrow = row+i
        for j in range(1,len(case[i])+1):
            sheet.cell(row = newrow,column= j).value = str(case[i][j-1])
    wb.save(filename)

if __name__ == '__main__':
    import os
    write_info(os.path.join(os.getcwd(),'test.xlsx'),['col1l,col12,col13','col2l,col22,col23,col24','col3l,col32,col33','col4l,col42,col43','','col6l,col62,col63'],10,6)
    write_result(os.path.join(os.getcwd(),'test.xlsx'),['col1l,col12,col13','col2l,col22,col23,col24','col3l,col32,col33','col4l,col42,col43','','col6l,col62,col63'])
    write_cell(os.path.join(os.getcwd(),'test.xlsx'),'test write cell3',3,3)
    print(read_info(os.path.join(os.getcwd(),'test.xlsx'),1,1))
    print(read_cell(os.path.join(os.getcwd(),'test.xlsx'),1,1))
